"""Rebuild GAMES_TRACKER.xlsx with remediation columns for the 2026-08 review response.

Carries every existing row forward untouched, appends the four remediation columns
requested by the client, keeps the local-only columns at the end so nothing is lost,
and appends the five new concepts.

    python scripts/build_remediation_tracker.py
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "GAMES_TRACKER.xlsx"

# Client-requested column order, then the four new columns, then local-only columns.
CLIENT_COLS = [
    "Game Concept",
    "Reference Game Link",
    "Financial Concept",
    "Game Name",
    "Game Name Bajaj",
    "Feedback on Reference Game",
    "Game Feedback",
    "Directory",
]
NEW_COLS = ["Remediation Type", "Effort", "Priority", "Status"]
TAIL_COLS = ["Type", "Dev Port", "Build Log"]
HEADERS = CLIENT_COLS + NEW_COLS + TAIL_COLS

# directory -> (Remediation Type, Effort, Priority, Status)
# Types: A asset swap | U UI rebuild | M mechanics redesign | F from scratch | B blocked
REMEDIATION = {
    # --- Tier 1: asset swap ------------------------------------------------
    "smart-match-3d": ("A — asset swap", "2d", "P1", "Phase 2 — proving trio"),
    "wealth-drop": ("A+U — asset swap + UI rebuild", "3d", "P1", "Phase 2 — proving trio"),
    "milestone-hopper": ("M+A — mechanics redesign + assets", "5d", "P1", "Phase 2 — proving trio"),
    "portfolio-fit": ("A — asset swap", "2d", "P2", "Phase 3 — stream A"),
    "risk-strike": ("A+U — asset swap + UI rebuild", "3d", "P2", "Phase 3 — stream A"),
    "steady-tower": ("A+U — asset swap + UI rebuild", "3d", "P2", "Phase 3 — stream A"),
    "spiral-sprint": ("A+U+M — assets, UI, longer run", "4d", "P2", "Phase 3 — stream A"),
    # --- Tier 2/3: mechanics ----------------------------------------------
    "swing-to-secure": ("M+U+A — mechanics, UI, assets", "6d", "P2", "Phase 3 — stream B"),
    "life-soar": ("M+A — mechanics redesign + assets", "6d", "P2", "Phase 3 — stream B"),
    "secure-journey": ("M+A — virus HP tiers, real powerups", "7d", "P2", "Phase 3 — stream B"),
    "coverage-archer": ("M+A — physics retune, single-player", "7d", "P2", "Phase 3 — stream C"),
    "tightrope-protection": ("M+A — restore wire, crows to virus", "6d", "P2", "Phase 3 — stream C"),
    # --- Tier 4: from scratch ---------------------------------------------
    "guardian-shelter": ("F — from scratch", "8d", "P2", "Phase 3 — stream C"),
    "risk-exit": ("F — from scratch", "7d", "P2", "Phase 3 — stream C"),
    "balance-block-journey": ("F — never built, new build", "7d", "P3", "Phase 4 — never built"),
    "shield-cascade": ("F — never built, new build", "8d", "P3", "Phase 4 — never built"),
    # --- Blocked -----------------------------------------------------------
    "goal-orbit": ("B — blocked", "—", "BLOCKED", 'Awaiting client call ("need to be discussed")'),
}

# Everything else was reviewed in the prior 2026-08-03 cycle and is already resolved,
# but still inherits the new style guide.
DEFAULT_REMEDIATION = (
    "A — style-guide conformance",
    "2d",
    "P4",
    "Phase 5 — conformance sweep",
)

# Two titles in the 2026-08 review have no row in the local tracker: the approved
# Bubble Shooter lives in the bajaj-game-store repo, and ripple-shield was deleted in
# the 2026-08-03 cycle before this feedback arrived. Both need to appear in the
# response, so they are added explicitly.
REVIEW_GAPS = [
    {
        "Game Concept": "Bubble shooter — launch and pop colour-matched bubbles",
        "Reference Game Link": "Deployed in bajaj-game-store (life-goals-bubble-shooter)",
        "Financial Concept": "Goal-linked play — clear the board, clear the goal",
        "Game Name": "Life Goals Bubble Shooter",
        "Game Name Bajaj": "Bubble Shooter",
        "Feedback on Reference Game": "APPROVED — the only title signed off in the 2026-08 review",
        "Game Feedback": (
            "Approved. Portfolio-wide rules still apply: strip instructional text, remove email "
            "field, no emoji, hazards become the green virus. Do not touch the game loop."
        ),
        "Directory": "life-goals-bubble-shooter (bajaj-game-store repo)",
        "Remediation Type": "A — portfolio rules only",
        "Effort": "1d",
        "Priority": "P4",
        "Status": "Approved — conformance sweep only",
        "Type": "Approved",
        "Dev Port": "",
        "Build Log": "Deployed in bajaj-game-store",
    },
    {
        "Game Concept": "One-tap chain reaction — a single tap triggers a spreading burst",
        "Reference Game Link": "Original concept",
        "Financial Concept": "One decision protects the whole chain",
        "Game Name": "Ripple Shield",
        "Game Name Bajaj": "Ripple Shield",
        "Feedback on Reference Game": '2026-08 review: "Design is too simple and basic, change it totally"',
        "Game Feedback": (
            "DIRECTORY DELETED 2026-08-03 (port 5046 retired) — the deletion predates this "
            "feedback. RECOMMEND: let the deletion stand and use the slot for a new concept. "
            "Rebuilding a title the client already rejected costs the same as a fresh concept "
            "with none of the baggage. Awaiting your call."
        ),
        "Directory": "ripple-shield (deleted)",
        "Remediation Type": "F — from scratch, or drop (recommended)",
        "Effort": "8d / 0d",
        "Priority": "BLOCKED",
        "Status": "Decision needed — rebuild or let deletion stand",
        "Type": "Deleted",
        "Dev Port": "5046 (retired)",
        "Build Log": "Deleted 2026-08-03",
    },
]

NEW_CONCEPTS = [
    {
        "Game Concept": "Magnet polarity steering — one tap flips attract/repel to steer a drifting orb",
        "Reference Game Link": "Original concept — from scratch",
        "Financial Concept": "The right pull at the right moment: cover attracts what you keep and repels what you don't",
        "Game Name": "Polarity",
        "Game Name Bajaj": "Pull & Protect",
        "Feedback on Reference Game": "Proposed 2026-08-22 — pending BajajLife sign-off",
        "Game Feedback": "",
        "Directory": "polarity-cover",
        "Remediation Type": "N — new concept",
        "Effort": "6d",
        "Priority": "P5",
        "Status": "Proposed — build after Phase 2 validates style guide",
        "Type": "New",
        "Dev Port": 5080,
        "Build Log": "Not started",
    },
    {
        "Game Concept": "Load-balancing beam — drag a counterweight to keep a pivoting beam level",
        "Reference Game Link": "Original concept — from scratch",
        "Financial Concept": "A portfolio is a beam: shocks land on one side, you rebalance rather than panic-sell",
        "Game Name": "Even Keel",
        "Game Name Bajaj": "Steady Balance",
        "Feedback on Reference Game": "Proposed 2026-08-22 — pending BajajLife sign-off",
        "Game Feedback": "",
        "Directory": "even-keel",
        "Remediation Type": "N — new concept",
        "Effort": "5d",
        "Priority": "P5",
        "Status": "Proposed — build after Phase 2 validates style guide",
        "Type": "New",
        "Dev Port": 5081,
        "Build Log": "Not started",
    },
    {
        "Game Concept": "Service-queue time management — tap a customer, tap the matching life-goal token",
        "Reference Game Link": "Original concept — from scratch",
        "Financial Concept": "Needs-based advice: the right cover for the right person, before they walk",
        "Game Name": "Front Desk",
        "Game Name Bajaj": "Advisor Rush",
        "Feedback on Reference Game": "Proposed 2026-08-22 — pending BajajLife sign-off",
        "Game Feedback": "COMPLIANCE: depicts advice — keep tokens abstract, route to legal before submission",
        "Directory": "front-desk",
        "Remediation Type": "N — new concept",
        "Effort": "7d",
        "Priority": "P5",
        "Status": "Proposed — needs legal review before build",
        "Type": "New",
        "Dev Port": 5082,
        "Build Log": "Not started",
    },
    {
        "Game Concept": "Gear-train assembly — drag gears onto pegs to drive an output wheel at a target ratio",
        "Reference Game Link": "Original concept — from scratch",
        "Financial Concept": "A plan only works if every part turns the next one",
        "Game Name": "Clockwork",
        "Game Name Bajaj": "Plan in Motion",
        "Feedback on Reference Game": "Proposed 2026-08-22 — pending BajajLife sign-off",
        "Game Feedback": "",
        "Directory": "clockwork-plan",
        "Remediation Type": "N — new concept",
        "Effort": "6d",
        "Priority": "P5",
        "Status": "Proposed — build after Phase 2 validates style guide",
        "Type": "New",
        "Dev Port": 5083,
        "Build Log": "Not started",
    },
    {
        "Game Concept": "Graph untangling — drag nodes until no connecting cord crosses another",
        "Reference Game Link": "Original concept — from scratch",
        "Financial Concept": "Tangled finances hide risk; straighten the lines and you can see what you own",
        "Game Name": "Untangle",
        "Game Name Bajaj": "Sort My Money",
        "Feedback on Reference Game": "Proposed 2026-08-22 — pending BajajLife sign-off",
        "Game Feedback": "",
        "Directory": "untangle",
        "Remediation Type": "N — new concept",
        "Effort": "5d",
        "Priority": "P5",
        "Status": "Proposed — cheapest new build, best single-glance differentiator",
        "Type": "New",
        "Dev Port": 5084,
        "Build Log": "Not started",
    },
]

PRIORITY_FILL = {
    "P1": "FF7A2F",
    "P2": "FFB800",
    "P3": "5CCBF5",
    "P4": "2F4A64",
    "P5": "7CD41F",
    "BLOCKED": "FF4D4D",
}


def main() -> None:
    wb = openpyxl.load_workbook(SRC)
    src = wb["Games Tracker"]

    old_headers = [c.value for c in src[1]]
    rows = [
        dict(zip(old_headers, [c.value for c in row]))
        for row in src.iter_rows(min_row=2)
        if any(c.value is not None for c in row)
    ]
    print(f"carried forward: {len(rows)} existing rows")

    # Idempotent: drop a previously generated sheet so re-runs replace rather than stack.
    if "Remediation Tracker" in wb.sheetnames:
        del wb["Remediation Tracker"]
    out = wb.create_sheet("Remediation Tracker", 0)
    out.append(HEADERS)

    def emit(record: dict) -> None:
        out.append([record.get(h, "") for h in HEADERS])

    for r in rows:
        directory = (r.get("Directory") or "").strip()
        rem_type, effort, priority, status = REMEDIATION.get(directory, DEFAULT_REMEDIATION)
        emit(
            {
                **{h: r.get(h, "") for h in CLIENT_COLS},
                "Remediation Type": rem_type,
                "Effort": effort,
                "Priority": priority,
                "Status": status,
                "Type": r.get("Type", ""),
                "Dev Port": r.get("Dev Port", ""),
                # The old "Status" column is a build log, not a remediation status.
                "Build Log": r.get("Status", ""),
            }
        )

    for gap in REVIEW_GAPS:
        emit(gap)
    print(f"appended: {len(REVIEW_GAPS)} review rows missing from the local tracker")

    for concept in NEW_CONCEPTS:
        emit(concept)
    print(f"appended: {len(NEW_CONCEPTS)} new concepts")

    # ---- formatting -------------------------------------------------------
    header_fill = PatternFill("solid", fgColor="00529B")
    for cell in out[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    out.row_dimensions[1].height = 32
    out.freeze_panes = "A2"
    out.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{out.max_row}"

    widths = {
        "Game Concept": 46,
        "Reference Game Link": 30,
        "Financial Concept": 40,
        "Game Name": 20,
        "Game Name Bajaj": 20,
        "Feedback on Reference Game": 38,
        "Game Feedback": 38,
        "Directory": 22,
        "Remediation Type": 34,
        "Effort": 8,
        "Priority": 10,
        "Status": 34,
        "Type": 9,
        "Dev Port": 10,
        "Build Log": 40,
    }
    for i, h in enumerate(HEADERS, start=1):
        out.column_dimensions[get_column_letter(i)].width = widths.get(h, 18)

    pri_col = HEADERS.index("Priority") + 1
    for row in out.iter_rows(min_row=2, min_col=pri_col, max_col=pri_col):
        cell = row[0]
        colour = PRIORITY_FILL.get(str(cell.value))
        if colour:
            cell.fill = PatternFill("solid", fgColor=colour)
            cell.font = Font(bold=True, color="FFFFFF" if cell.value != "P4" else "FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    for row in out.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top", wrap_text=True, horizontal=cell.alignment.horizontal
            )

    # ---- keep the do-not-repeat catalog current ---------------------------
    catalog = wb["Do-Not-Repeat Catalog"]
    label = "New concepts (2026-08-22, pending sign-off)"
    already = any(row[0].value == label for row in catalog.iter_rows(min_col=1, max_col=1))
    if not already:
        catalog.append(
            [
                label,
                "polarity-cover (magnet attract/repel steering), even-keel (load-balancing beam), "
                "front-desk (service-queue time management), clockwork-plan (gear-train assembly), "
                "untangle (graph planarity). Mechanic families deliberately chosen as unused across "
                "both this repo and bajaj-game-store.",
            ]
        )

    wb.save(SRC)
    print(f"wrote {SRC} — {out.max_row - 1} rows, {len(HEADERS)} columns")


if __name__ == "__main__":
    main()
