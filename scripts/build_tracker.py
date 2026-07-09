"""Regenerate GAMES_TRACKER.xlsx from scripts/games-manifest.json.

Usage:  python scripts/build_tracker.py [--status-file scripts/build-status.json]
The optional status file maps directory name -> status string (e.g. "Done", "Build failed").
"""
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
MANIFEST = ROOT / "scripts" / "games-manifest.json"
OUT = ROOT / "GAMES_TRACKER.xlsx"

HEADERS = [
    "Game Concept", "Reference Game Link", "Financial Concept", "Game Name",
    "Game Name Bajaj", "Feedback on Reference Game", "Game Feedback",
    "Directory", "Type", "Status", "Dev Port",
]

BLUE = "003DA6"
ORANGE = "F26522"

CATALOG_NOTE = [
    ("This repo (existing)", "coverage-archer (archery), life-goals-bubble-shooter (bubble shooter — GOLD STANDARD), "
     "stackibility-stack (stacking), tightrope-protection (wire balance), retire-rich-clicker (clicker), "
     "edurise-jumper (vertical jumper), tax-save-maze (maze), she-shield-protector (catcher), "
     "safe-stride-balancer (balance)"),
    ("bajaj-game-store (do not repeat)", "snake, hangman, word scramble, match-3 x2, tetris, fruit-slice, "
     "runners/climbers x4, top-down shooter, bomberman, galaga, sudoku x2, whack-a-mole x2, brick-breaker, "
     "jigsaw x2, minesweeper, racing, tube-sorting, quiz x3, bubble shooter, peg solitaire, stacking, "
     "memory-flip, tower defense, snakes & ladders, arcade dodger"),
    ("Dropped by feedback", "Wealth Current (physics orb), Path to Legacy (line drawing), "
     "Launch to Protection (hedgehog launch), Secure Foundations (falling sand)"),
]


def main() -> None:
    status_map = {}
    args = sys.argv[1:]
    if len(args) == 2 and args[0] == "--status-file":
        p = Path(args[1])
        if p.exists():
            status_map = json.loads(p.read_text(encoding="utf-8"))

    data = json.loads(MANIFEST.read_text(encoding="utf-8"))
    wb = Workbook()

    ws = wb.active
    ws.title = "Games Tracker"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    def add_row(g, gtype):
        status = status_map.get(g["dir"], "In Progress")
        ws.append([
            g["concept"], g.get("reference", ""), g["financialConcept"], g.get("gameName", ""),
            g["bajajName"], g.get("feedback", ""), "",
            g["dir"], gtype, status, g.get("port") or "",
        ])

    for g in data["newGames"]:
        add_row(g, "New")
    for g in data["revamps"]:
        add_row(g, "Revamp")
    ws.append([
        "Bubble shooter (Puzzle Bobble style)", "", "Life goals via goal-linked play",
        "Life Goals Bubble Shooter", "Life Goals Bubble Shooter", "", "Gold standard — untouched",
        "life-goals-bubble-shooter", "Standard", "Done (reference)", 5018,
    ])

    widths = [34, 40, 38, 24, 22, 44, 30, 24, 10, 14, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        t = row[8].value
        if t == "Revamp":
            row[8].fill = PatternFill("solid", fgColor="FDE9DD")
        elif t == "New":
            row[8].fill = PatternFill("solid", fgColor="E3ECFA")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

    ws2 = wb.create_sheet("Do-Not-Repeat Catalog")
    ws2.append(["Source", "Mechanics / games already covered"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=ORANGE)
    for src, note in CATALOG_NOTE:
        ws2.append([src, note])
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 120
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(OUT)
    print(f"Wrote {OUT} ({ws.max_row - 1} game rows)")


if __name__ == "__main__":
    main()
